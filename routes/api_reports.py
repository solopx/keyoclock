"""
routes/api_reports.py
API de relatórios: dados agregados e exportação XLSX/PDF.
"""
import io
import logging
from datetime import datetime
from flask import Blueprint, jsonify, send_file, request as flask_request
from database.db import qall, qone, get_enc_key, decrypt_val
from database.auth import login_required
from database.helpers import lic_status
from database import ratelimit

reports_bp = Blueprint('reports', __name__)
_logger = logging.getLogger(__name__)


@reports_bp.route('/api/reports/data')
@login_required
def reports_data():
    # Reutiliza o mesmo helper usado pelo export PDF/XLSX (elimina query duplicada)
    lics = _query_licenses_full()

    by_type     = {}
    by_division = {}
    by_status   = dict.fromkeys(['valid', 'expired', 'perpetual', 'critical', 'warning', 'soon', 'unknown'], 0)

    # breakdown por grupo: { div_name: { status: qty, ... } }
    by_division_status = {}

    for l in lics:
        qty    = l['quantity'] or 1
        status, days = lic_status(l)
        by_type[l['license_type']] = by_type.get(l['license_type'], 0) + qty
        by_division[l['dname']]    = by_division.get(l['dname'], 0) + qty
        by_status[status]          = by_status.get(status, 0) + qty

        # acumula status dentro de cada grupo
        dname = l['dname']
        if dname not in by_division_status:
            by_division_status[dname] = dict.fromkeys(
                ['valid','expired','perpetual','critical','warning','soon','unknown'], 0)
        by_division_status[dname][status] = by_division_status[dname].get(status, 0) + qty

    total_qty = sum((l['quantity'] or 1) for l in lics)

    # 5 licenças que vencem mais cedo (excluindo perpétuas e já expiradas)
    upcoming = sorted(
        [l for l in lics if l.get('end_date') and not l['is_perpetual']],
        key=lambda l: l['end_date']
    )[:5]
    upcoming_out = []
    for l in upcoming:
        status, days = lic_status(l)
        upcoming_out.append({
            'item_id':       l['item_id'],
            'item_name':     l.get('iname', ''),
            'div_name':      l['dname'],
            'license_type':  l['license_type'],
            'quantity':      l['quantity'] or 1,
            'end_date':      l['end_date'],
            'status':        status,
            'days':          days,
        })

    return jsonify({
        'by_type':            by_type,
        'by_status':          by_status,
        'by_division':        by_division,
        'by_division_status': by_division_status,
        'upcoming_expiry':    upcoming_out,
        'total_items':        qone("SELECT COUNT(*) c FROM inventory_item WHERE deleted_at IS NULL")['c'],
        'total_licenses':     total_qty,
        'total_entries':      len(lics),
        'total_divisions':    qone("SELECT COUNT(*) c FROM division      WHERE deleted_at IS NULL")['c'],
        'total_lists':        qone("SELECT COUNT(*) c FROM device_list   WHERE deleted_at IS NULL")['c'],
    })


@reports_bp.route('/api/reports/export/xlsx')
@login_required
def export_xlsx():
    ip = flask_request.remote_addr or '0.0.0.0'
    if not ratelimit.check(f'report_export:{ip}', max_calls=5, window=60):
        return jsonify({'error': 'Muitas requisições. Aguarde um momento.'}), 429
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl não instalado. Execute: pip install openpyxl'}), 500

    status_filter = flask_request.args.get('filter', 'all')

    wb = openpyxl.Workbook()
    _build_licenses_sheet(wb.active, Font, PatternFill, Alignment, status_filter=status_filter)
    if status_filter != 'all':
        _build_licenses_sheet(wb.create_sheet("Todas as Licenças"), Font, PatternFill, Alignment)
    _build_inventory_sheet(wb.create_sheet("Inventário"), Font, PatternFill, Alignment)
    _build_contracts_sheet(wb.create_sheet("Por Contrato"), Font, PatternFill, Alignment)
    _build_groups_sheet(wb.create_sheet("Por Grupo"), Font, PatternFill, Alignment)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"licencas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route('/api/reports/export/pdf')
@login_required
def export_pdf():
    ip = flask_request.remote_addr or '0.0.0.0'
    if not ratelimit.check(f'report_export:{ip}', max_calls=5, window=60):
        return jsonify({'error': 'Muitas requisições. Aguarde um momento.'}), 429
    from flask import request
    try:
        tab           = request.args.get('tab', 'overview')
        status_filter = request.args.get('filter', 'all')
        pdf_bytes, filename = generate_pdf_bytes(tab, status_filter)
        buf = io.BytesIO(pdf_bytes)
        return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)
    except ImportError:
        return jsonify({'error': 'reportlab não instalado. Execute: pip install reportlab'}), 500
    except Exception as e:
        _logger.error('[export_pdf] %s', e, exc_info=True)
        return jsonify({'error': 'Erro ao gerar PDF. Verifique os logs do servidor.'}), 500


def generate_pdf_bytes(tab='overview', status_filter='all'):
    """Gera o PDF de relatório e retorna (bytes, filename). Raises ImportError se reportlab ausente."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_LEFT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1*cm
    )
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    sub_style   = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9,
                                 textColor=colors.HexColor('#64748b'), spaceAfter=6)
    h2_style    = ParagraphStyle('h2', parent=styles['Heading2'], fontSize=11,
                                 textColor=colors.HexColor('#1e3a5f'), spaceBefore=14, spaceAfter=6)

    HDR  = colors.HexColor('#1e3a5f')
    ALT  = colors.HexColor('#f0f4f8')
    WHT  = colors.white
    GRID = colors.HexColor('#cbd5e1')

    STATUS_COLORS = {
        'valid':     colors.HexColor('#10b981'),
        'perpetual': colors.HexColor('#06b6d4'),
        'expired':   colors.HexColor('#9333ea'),  # roxo
        'critical':  colors.HexColor('#ef4444'),  # vermelho
        'warning':   colors.HexColor('#f97316'),  # laranja
        'soon':      colors.HexColor('#f59e0b'),  # amarelo
        'unknown':   colors.HexColor('#94a3b8'),
    }
    STATUS_LABELS = {'valid':'Saudável','perpetual':'Vitalícia','expired':'Expirada',
                     'critical':'Crítico (≤30d)','warning':'Aviso (≤60d)','soon':'Atenção (≤90d)','unknown':'?'}

    def make_table(rows, col_widths=None):
        t = Table(rows, repeatRows=1, colWidths=col_widths)
        style = [
            ('BACKGROUND', (0,0), (-1,0), HDR),
            ('TEXTCOLOR',  (0,0), (-1,0), WHT),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 7),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHT, ALT]),
            ('GRID',  (0,0), (-1,-1), 0.3, GRID),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',(0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,1), (2,-1), 'LEFT'),   # primeiras colunas alinhadas à esquerda
        ]
        t.setStyle(TableStyle(style))
        return t

    story = []
    ts    = datetime.now().strftime('%d/%m/%Y %H:%M')

    # ── ABA: VISÃO GERAL ──────────────────────────────────────────────────────
    if tab == 'overview':
        story.append(Paragraph(f'Relatório Geral de Licenças', title_style))
        story.append(Paragraph(f'Gerado em {ts}', sub_style))
        story.append(Spacer(1, 0.3*cm))

        # Resumo por status
        lics = _query_licenses_full()
        by_status   = {}
        by_type     = {}
        by_division = {}
        total_qty   = 0
        for l in lics:
            qty  = l['quantity'] or 1
            s, _ = lic_status(l)
            by_status[s]              = by_status.get(s,0) + qty
            by_type[l['license_type']] = by_type.get(l['license_type'],0) + qty
            by_division[l['dname']]   = by_division.get(l['dname'],0) + qty
            total_qty += qty

        story.append(Paragraph('Resumo por Status', h2_style))
        rows = [['Status', 'Quantidade', '% do Total']]
        for s, qty in sorted(by_status.items(), key=lambda x: -x[1]):
            if qty == 0: continue
            pct = f"{round(qty/total_qty*100)}%" if total_qty else '0%'
            rows.append([STATUS_LABELS.get(s, s), str(qty), pct])
        rows.append(['TOTAL', str(total_qty), '100%'])
        story.append(make_table(rows, col_widths=[8*cm, 4*cm, 3*cm]))
        story.append(Spacer(1, 0.4*cm))

        story.append(Paragraph('Resumo por Tipo', h2_style))
        rows = [['Tipo', 'Quantidade', '% do Total']]
        for tp, qty in sorted(by_type.items(), key=lambda x: -x[1]):
            pct = f"{round(qty/total_qty*100)}%" if total_qty else '0%'
            rows.append([tp.capitalize(), str(qty), pct])
        story.append(make_table(rows, col_widths=[8*cm, 4*cm, 3*cm]))
        story.append(Spacer(1, 0.4*cm))

        story.append(Paragraph('Resumo por Grupo', h2_style))
        rows = [['Grupo', 'Quantidade', '% do Total']]
        for div, qty in sorted(by_division.items(), key=lambda x: -x[1]):
            pct = f"{round(qty/total_qty*100)}%" if total_qty else '0%'
            rows.append([div, str(qty), pct])
        story.append(make_table(rows, col_widths=[8*cm, 4*cm, 3*cm]))
        story.append(Spacer(1, 0.4*cm))

        story.append(Paragraph('Resumo por Contrato', h2_style))
        key = get_enc_key()
        _S_ORDER = {'expired':0,'critical':1,'warning':2,'soon':3,'valid':4,'perpetual':5,'unknown':6}
        ctr_summary = {}
        for l in lics:
            raw  = l.get('contract') or ''
            name = decrypt_val(key, raw) if raw else ''
            grp  = name if name else '(Sem contrato)'
            s, _ = lic_status(l)
            qty  = l['quantity'] or 1
            if grp not in ctr_summary:
                ctr_summary[grp] = {'license_count': 0, 'total_qty': 0, 'worst_order': 99, 'worst_status': 'unknown'}
            c = ctr_summary[grp]
            c['license_count'] += 1
            c['total_qty']     += qty
            order = _S_ORDER.get(s, 99)
            if order < c['worst_order']:
                c['worst_order']  = order
                c['worst_status'] = s
        def _ctr_sort(kv):
            return (1 if kv[0] == '(Sem contrato)' else 0, _S_ORDER.get(kv[1]['worst_status'], 99), kv[0].lower())
        rows = [['Contrato', 'Registros', 'Qtd. Total', 'Pior Status']]
        sorted_ctrs = sorted(ctr_summary.items(), key=_ctr_sort)
        for cname, cd in sorted_ctrs[:20]:
            rows.append([cname, str(cd['license_count']), str(cd['total_qty']),
                         STATUS_LABELS.get(cd['worst_status'], cd['worst_status'])])
        if len(ctr_summary) > 20:
            rows.append([f'... e mais {len(ctr_summary)-20} contratos', '', '', ''])
        story.append(make_table(rows, col_widths=[8*cm, 3*cm, 3*cm, 3*cm]))

    # ── ABA: POR VALIDADE ─────────────────────────────────────────────────────
    elif tab == 'validity':
        label_map = {'all':'Todas','valid':'Saudáveis','perpetual':'Vitalícias',
                     'expired':'Expiradas','critical':'Crítico (≤30d)',
                     'warning':'Aviso (≤60d)','soon':'Atenção (≤90d)'}
        label = label_map.get(status_filter, status_filter)
        story.append(Paragraph(f'Relatório por Validade — {label}', title_style))
        story.append(Paragraph(f'Gerado em {ts}', sub_style))
        story.append(Spacer(1, 0.3*cm))

        lics = _query_licenses_full()
        if status_filter != 'all':
            lics = [l for l in lics if lic_status(l)[0] == status_filter]

        # Ordena: expiradas/críticas primeiro, depois por dias
        def sort_key(l):
            order = {'expired':0,'critical':1,'warning':2,'soon':3,'valid':4,'perpetual':5,'unknown':6}
            s, days = lic_status(l)
            return (order.get(s,7), days if days is not None else 9999)
        lics = sorted(lics, key=sort_key)

        rows = [['Grupo', 'Lista', 'Item', 'Tipo', 'Qtd', 'Contrato', 'Início', 'Vencimento', 'Status', 'Dias']]
        style_cmds = [
            ('BACKGROUND', (0,0), (-1,0), HDR),
            ('TEXTCOLOR',  (0,0), (-1,0), WHT),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 7),
            ('GRID',  (0,0), (-1,-1), 0.3, GRID),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',(0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,1), (3,-1), 'LEFT'),
        ]
        key = get_enc_key()
        for i, l in enumerate(lics, 1):
            s, days = lic_status(l)
            days_str = str(days) + 'd' if days is not None else '∞'
            rows.append([
                l['dname'][:16], l['lname'][:16], l['iname'][:20],
                l['license_type'], str(l['quantity'] or 1),
                decrypt_val(key, l.get('contract') or '') or '—',
                l['start_date'] or '—', l['end_date'] or '—',
                STATUS_LABELS.get(s, s), days_str,
            ])
            # colorir célula de status
            sc = STATUS_COLORS.get(s, colors.HexColor('#94a3b8'))
            style_cmds.append(('TEXTCOLOR', (8,i), (8,i), sc))
            style_cmds.append(('FONTNAME',  (8,i), (8,i), 'Helvetica-Bold'))
            if s == 'expired':
                style_cmds.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#faf0ff')))
            elif s == 'critical':
                style_cmds.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#fff1f2')))
            elif s == 'warning':
                style_cmds.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#fff7ed')))
            elif s == 'soon':
                style_cmds.append(('BACKGROUND', (0,i), (-1,i), colors.HexColor('#fffbeb')))

        if len(rows) == 1:
            story.append(Paragraph('Nenhuma licença encontrada para este filtro.', sub_style))
        else:
            t = Table(rows, repeatRows=1,
                      colWidths=[3.2*cm,3.2*cm,4*cm,2.5*cm,1.2*cm,3*cm,2.2*cm,2.2*cm,2*cm,1.5*cm])
            t.setStyle(TableStyle(style_cmds))
            story.append(t)

    # ── ABA: POR GRUPO ────────────────────────────────────────────────────────
    elif tab == 'divisions':
        story.append(Paragraph('Relatório por Grupo', title_style))
        story.append(Paragraph(f'Gerado em {ts}', sub_style))
        story.append(Spacer(1, 0.3*cm))

        lics = _query_licenses_full()
        # Agrupar por divisão
        from collections import defaultdict, OrderedDict
        divs = OrderedDict()
        for l in lics:
            dn = l['dname']
            if dn not in divs:
                divs[dn] = []
            divs[dn].append(l)

        total_geral = sum((l['quantity'] or 1) for l in lics)
        key = get_enc_key()

        for div_name, div_lics in divs.items():
            total_div = sum((l['quantity'] or 1) for l in div_lics)
            pct_div   = f"{round(total_div/total_geral*100)}%" if total_geral else '0%'
            story.append(Paragraph(f'Grupo: {div_name}  •  {total_div} licenças  ({pct_div} do total)', h2_style))

            rows = [['Lista', 'Item', 'Tipo', 'Qtd', 'Contrato', 'Início', 'Vencimento', 'Status', 'Dias']]
            style_cmds = [
                ('BACKGROUND', (0,0), (-1,0), HDR),
                ('TEXTCOLOR',  (0,0), (-1,0), WHT),
                ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0,0), (-1,-1), 7),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHT, ALT]),
                ('GRID',  (0,0), (-1,-1), 0.3, GRID),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',(0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,1), (2,-1), 'LEFT'),
            ]
            for i, l in enumerate(div_lics, 1):
                s, days = lic_status(l)
                rows.append([
                    l['lname'][:18], l['iname'][:22], l['license_type'],
                    str(l['quantity'] or 1), decrypt_val(key, l.get('contract') or '') or '—',
                    l['start_date'] or '—', l['end_date'] or '—',
                    STATUS_LABELS.get(s,s),
                    str(days)+'d' if days is not None else '∞',
                ])
                sc = STATUS_COLORS.get(s, colors.HexColor('#94a3b8'))
                style_cmds.append(('TEXTCOLOR', (7,i), (7,i), sc))
                style_cmds.append(('FONTNAME',  (7,i), (7,i), 'Helvetica-Bold'))

            t = Table(rows, repeatRows=1,
                      colWidths=[3.5*cm,5*cm,2.8*cm,1.5*cm,3*cm,2.2*cm,2.2*cm,2.2*cm,1.6*cm])
            t.setStyle(TableStyle(style_cmds))
            story.append(t)
            story.append(Spacer(1, 0.5*cm))

    # ── ABA: POR CONTRATO ─────────────────────────────────────────────────────
    elif tab == 'contracts':
        story.append(Paragraph('Relatorio por Contrato', title_style))
        story.append(Paragraph(f'Gerado em {ts}', sub_style))
        story.append(Spacer(1, 0.3*cm))

        lics = _query_licenses_full()
        key  = get_enc_key()

        _S_ORDER = {'expired':0,'critical':1,'warning':2,'soon':3,'valid':4,'perpetual':5,'unknown':6}
        ctr_map  = {}
        for l in lics:
            raw   = l.get('contract') or ''
            name  = decrypt_val(key, raw) if raw else ''
            group = name if name else '(Sem contrato)'
            s, days = lic_status(l)
            qty     = l['quantity'] or 1
            if group not in ctr_map:
                ctr_map[group] = {'worst_order': 99, 'worst_status': 'unknown', 'items': []}
            c = ctr_map[group]
            order = _S_ORDER.get(s, 99)
            if order < c['worst_order']:
                c['worst_order']  = order
                c['worst_status'] = s
            c['items'].append((l, s, days, qty))

        def _ctr_key(kv):
            return (1 if kv[0] == '(Sem contrato)' else 0, kv[0].lower())

        for cname, cdata in sorted(ctr_map.items(), key=_ctr_key):
            story.append(Paragraph(f'Contrato: {cname}', h2_style))
            rows = [['Item', 'Localizacao', 'Tipo', 'Qtd', 'Inicio', 'Vencimento', 'Status', 'Dias']]
            style_cmds = [
                ('BACKGROUND', (0,0), (-1,0), HDR),
                ('TEXTCOLOR',  (0,0), (-1,0), WHT),
                ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0,0), (-1,-1), 7),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHT, ALT]),
                ('GRID',  (0,0), (-1,-1), 0.3, GRID),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',(0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,1), (1,-1), 'LEFT'),
            ]
            sorted_items = sorted(cdata['items'],
                                  key=lambda x: (_S_ORDER.get(x[1], 99), x[0].get('end_date') or '9999'))
            for i, (l, s, days, qty) in enumerate(sorted_items, 1):
                location = f"{l.get('dname','')[:14]} > {l.get('lname','')[:14]}"
                perp     = bool(l.get('is_perpetual'))
                days_str = ('--' if perp else (str(days) + 'd' if days is not None else '?'))
                rows.append([
                    l.get('iname', '')[:22],
                    location,
                    l.get('license_type', ''),
                    str(qty),
                    l.get('start_date') or '-',
                    'Vitalicia' if perp else (l.get('end_date') or '-'),
                    STATUS_LABELS.get(s, s),
                    days_str,
                ])
                sc = STATUS_COLORS.get(s, colors.HexColor('#94a3b8'))
                style_cmds.append(('TEXTCOLOR', (6, i), (6, i), sc))
                style_cmds.append(('FONTNAME',  (6, i), (6, i), 'Helvetica-Bold'))
            t = Table(rows, repeatRows=1,
                      colWidths=[4*cm, 4.5*cm, 2.5*cm, 1.2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 1.5*cm])
            t.setStyle(TableStyle(style_cmds))
            story.append(t)
            story.append(Spacer(1, 0.5*cm))

    if not story:
        story.append(Paragraph('Exportacao nao disponivel para esta aba.', styles['Normal']))
    doc.build(story)
    buf.seek(0)
    tab_suffix = {'overview':'geral','validity':f'validade_{status_filter}','divisions':'grupos','contracts':'contratos'}.get(tab,'relatorio')
    filename = f"keyoclock_{tab_suffix}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buf.read(), filename


@reports_bp.route('/api/reports/contracts')
@login_required
def reports_contracts():
    lics = _query_licenses_full()
    key  = get_enc_key()

    STATUS_ORDER = {'expired':0,'critical':1,'warning':2,'soon':3,'valid':4,'perpetual':5,'unknown':6}
    contracts    = {}

    for l in lics:
        raw = l.get('contract') or ''
        name = decrypt_val(key, raw) if raw else ''
        group = name if name else '(Sem contrato)'

        status, days = lic_status(l)
        qty          = l['quantity'] or 1

        if group not in contracts:
            contracts[group] = {
                'name':          group,
                'license_count': 0,
                'total_qty':     0,
                'earliest_end':  None,
                'worst_order':   99,
                'worst_status':  'unknown',
                'items':         [],
            }

        c = contracts[group]
        c['license_count'] += 1
        c['total_qty']     += qty

        if not l['is_perpetual'] and l.get('end_date'):
            if c['earliest_end'] is None or l['end_date'] < c['earliest_end']:
                c['earliest_end'] = l['end_date']

        order = STATUS_ORDER.get(status, 99)
        if order < c['worst_order']:
            c['worst_order']  = order
            c['worst_status'] = status

        c['items'].append({
            'item_name':    l.get('iname', ''),
            'division_name':l.get('dname', ''),
            'list_name':    l.get('lname', ''),
            'license_type': l.get('license_type', ''),
            'quantity':     qty,
            'start_date':   l.get('start_date') or '',
            'end_date':     l.get('end_date') or '',
            'status':       status,
            'days':         days,
            'is_perpetual': bool(l.get('is_perpetual')),
        })

    def _sort_key(c):
        return (1 if c['name'] == '(Sem contrato)' else 0, c['name'].lower())

    result = []
    for c in sorted(contracts.values(), key=_sort_key):
        c['items'].sort(key=lambda i: (STATUS_ORDER.get(i['status'], 99), i['end_date'] or '9999'))
        del c['worst_order']
        result.append(c)

    return jsonify(result)


# ─── HELPERS PRIVADOS ─────────────────────────────────────────────────────────

def query_licenses_full():
    return _query_licenses_full()


def _query_licenses_full():
    return qall("""
        SELECT l.*, i.name iname, i.model imodel, dl.name lname, dv.name dname
        FROM license l
        JOIN inventory_item i  ON l.item_id      = i.id
        JOIN device_list   dl  ON i.list_id       = dl.id
        JOIN division      dv  ON dl.division_id  = dv.id
        WHERE l.deleted_at  IS NULL
          AND i.deleted_at  IS NULL
          AND dl.deleted_at IS NULL
          AND dv.deleted_at IS NULL
        ORDER BY dv.name, dl.name, i.name, l.end_date
    """)


def _build_licenses_sheet(ws, Font, PatternFill, Alignment, status_filter='all'):
    _STATUS_LABELS_FILTER = {
        'expired': 'Expiradas', 'critical': 'Crítico', 'warning': 'Aviso',
        'soon': 'Atenção', 'valid': 'Saudáveis', 'perpetual': 'Vitalícias',
    }
    ws.title = "Licenças" if status_filter == 'all' else f"Licenças — {_STATUS_LABELS_FILTER.get(status_filter, status_filter)}"
    headers  = ['Grupo', 'Lista', 'Item', 'Modelo', 'Tipo', 'Qtd', 'Contrato',
                 'Vitalícia', 'Início', 'Fim', 'Status', 'Dias', 'Notas']
    hfill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment = hfill, hfont, Alignment(horizontal='center')

    key  = get_enc_key()
    lics = _query_licenses_full()
    if status_filter != 'all':
        lics = [l for l in lics if lic_status(l)[0] == status_filter]
    for l in lics:
        status, days = lic_status(l)
        ws.append([
            l['dname'], l['lname'], l['iname'], l.get('imodel', ''),
            l['license_type'], l['quantity'] or 1,
            decrypt_val(key, l.get('contract') or ''),
            'Sim' if l['is_perpetual'] else 'Não',
            l['start_date'] or '', l['end_date'] or '',
            status, days if days is not None else 'N/A',
            l.get('notes', ''),
        ])

    _auto_width(ws)


def _build_inventory_sheet(ws, Font, PatternFill, Alignment):
    headers = ['Divisão', 'Lista', 'Item', 'Modelo', 'Fabricante', 'Fornecedor', 'Descrição', 'Qtd Licenças']
    hfill   = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    hfont   = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font = hfill, hfont

    items = qall("""
        SELECT i.*, dl.name lname, dv.name dname,
               COALESCE((SELECT SUM(quantity) FROM license
                         WHERE item_id = i.id AND deleted_at IS NULL), 0) lcount
        FROM inventory_item i
        JOIN device_list dl ON i.list_id      = dl.id
        JOIN division   dv ON dl.division_id  = dv.id
        WHERE i.deleted_at  IS NULL
          AND dl.deleted_at IS NULL
          AND dv.deleted_at IS NULL
        ORDER BY dv.name, dl.name, i.name
    """)
    for it in items:
        ws.append([
            it['dname'], it['lname'], it['name'], it.get('model', ''),
            it.get('manufacturer', ''), it.get('supplier', ''),
            it.get('description', ''), it['lcount'],
        ])

    _auto_width(ws)


def _build_contracts_sheet(ws, Font, PatternFill, Alignment):
    ws.title = "Por Contrato"
    headers  = ['Item', 'Modelo', 'Localização', 'Tipo', 'Qtd', 'Início', 'Vencimento', 'Status', 'Dias']
    hfill    = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    hfont    = Font(color='FFFFFF', bold=True)
    gfill    = PatternFill(start_color='e8f0fe', end_color='e8f0fe', fill_type='solid')
    gfont    = Font(bold=True, color='1e3a5f')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment = hfill, hfont, Alignment(horizontal='center')

    _S_ORDER = {'expired':0,'critical':1,'warning':2,'soon':3,'valid':4,'perpetual':5,'unknown':6}
    key  = get_enc_key()
    lics = _query_licenses_full()

    ctr_map = {}
    for l in lics:
        raw   = l.get('contract') or ''
        name  = decrypt_val(key, raw) if raw else ''
        group = name if name else '(Sem contrato)'
        s, days = lic_status(l)
        if group not in ctr_map:
            ctr_map[group] = []
        ctr_map[group].append((l, s, days))

    def _ctr_key(k):
        return (1 if k == '(Sem contrato)' else 0, k.lower())

    row_num = 2
    for cname in sorted(ctr_map.keys(), key=_ctr_key):
        items = sorted(ctr_map[cname],
                       key=lambda x: (_S_ORDER.get(x[1], 99), x[0].get('end_date') or '9999'))
        g = ws.cell(row=row_num, column=1, value=f'Contrato: {cname}')
        g.font, g.fill = gfont, gfill
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        row_num += 1
        for l, s, days in items:
            perp = bool(l.get('is_perpetual'))
            for col, val in enumerate([
                l.get('iname', ''), l.get('imodel', ''),
                f"{l.get('dname','')} › {l.get('lname','')}",
                l.get('license_type', ''), l['quantity'] or 1,
                l.get('start_date') or '',
                '—' if perp else (l.get('end_date') or ''),
                s, days if days is not None else 'N/A',
            ], 1):
                ws.cell(row=row_num, column=col, value=val)
            row_num += 1

    _auto_width(ws)


def _build_groups_sheet(ws, Font, PatternFill, Alignment):
    ws.title = "Por Grupo"
    headers  = ['Lista', 'Item', 'Modelo', 'Tipo', 'Qtd', 'Contrato', 'Início', 'Vencimento', 'Status', 'Dias']
    hfill    = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    hfont    = Font(color='FFFFFF', bold=True)
    gfill    = PatternFill(start_color='e8f0fe', end_color='e8f0fe', fill_type='solid')
    gfont    = Font(bold=True, color='1e3a5f')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment = hfill, hfont, Alignment(horizontal='center')

    _S_ORDER = {'expired':0,'critical':1,'warning':2,'soon':3,'valid':4,'perpetual':5,'unknown':6}
    key  = get_enc_key()
    lics = _query_licenses_full()

    grp_map = {}
    for l in lics:
        dname = l['dname']
        if dname not in grp_map:
            grp_map[dname] = []
        grp_map[dname].append(l)

    row_num = 2
    for dname in sorted(grp_map.keys()):
        items = sorted(grp_map[dname],
                       key=lambda l: (_S_ORDER.get(lic_status(l)[0], 99), l.get('end_date') or '9999'))
        g = ws.cell(row=row_num, column=1, value=f'Grupo: {dname}')
        g.font, g.fill = gfont, gfill
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        row_num += 1
        for l in items:
            s, days = lic_status(l)
            for col, val in enumerate([
                l.get('lname', ''), l.get('iname', ''), l.get('imodel', ''),
                l.get('license_type', ''), l['quantity'] or 1,
                decrypt_val(key, l.get('contract') or ''),
                l.get('start_date') or '',
                '—' if l.get('is_perpetual') else (l.get('end_date') or ''),
                s, days if days is not None else 'N/A',
            ], 1):
                ws.cell(row=row_num, column=col, value=val)
            row_num += 1

    _auto_width(ws)


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)
